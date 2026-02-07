from io import BytesIO
from openpyxl import load_workbook
from src.processing.extractors.base import BaseExtractor, ExtractedDocument


class ExcelExtractor(BaseExtractor):
    MAX_ROWS = 1000  # Limit rows to prevent memory issues

    def extract(self, file_content: bytes) -> ExtractedDocument:
        """Extract text from an Excel file using openpyxl."""
        wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)

        all_text = []
        sheet_count = 0

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_count += 1
            sheet_text = [f"## Sheet: {sheet_name}"]

            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                if row_count >= self.MAX_ROWS:
                    sheet_text.append(f"... (truncated at {self.MAX_ROWS} rows)")
                    break

                # Filter out None values and convert to strings
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(v.strip() for v in row_values):  # Skip empty rows
                    sheet_text.append(" | ".join(row_values))
                    row_count += 1

            all_text.append("\n".join(sheet_text))

        wb.close()
        content = "\n\n".join(all_text)

        return ExtractedDocument(
            content=content,
            page_count=sheet_count,
            metadata={"extractor": "openpyxl", "sheet_count": sheet_count},
            method="native_excel",
        )
